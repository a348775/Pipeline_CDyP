from confluent_kafka import Consumer
import json
import cx_Oracle
import pandas as pd
from datetime import datetime
import time
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill

cx_Oracle.init_oracle_client(lib_dir=r"C:\oracle\instantclient\instantclient_23_6")

def parse_date(date_string):
    try:
        return datetime.strptime(date_string, "%Y-%m-%d")
    except ValueError:
        print(f"Fecha inválida: {date_string}")
        return None

# Configuración del consumidor Kafka
consumer_config = {
    'bootstrap.servers': 'localhost:9092',
    'group.id': 'multi-api-consumer-group',
    'auto.offset.reset': 'earliest'
}
consumer = Consumer(consumer_config)
topics = ["nasa-data", "weather-data"]
consumer.subscribe(topics)

# Conexión a Oracle
dsn = cx_Oracle.makedsn("localhost", 1521, service_name="XE")
connection = cx_Oracle.connect(user="computo", password="computo", dsn=dsn)

# Función para insertar datos de NASA
def insert_nasa_data(data):
    cursor = connection.cursor()
    query = """
        INSERT INTO nasa_data ("NASA_DATE", title, explanation, media_type, url)
        VALUES (TO_DATE(:nasa_date, 'YYYY-MM-DD'), :title, :explanation, :media_type, :url)
    """
    # Validar y transformar datos
    parsed_date = parse_date(data.get('date'))  # Convertir la fecha
    if not parsed_date:
        print(f"Fecha inválida en datos: {data}")
        return
    
    cursor.execute(query, {
        'nasa_date': parsed_date,  # Cambié ':date' por ':nasa_date'
        'title': data.get('title', "Sin título"),  # Valor predeterminado
        'explanation': data.get('explanation', ""),
        'media_type': data.get('media_type', ""),
        'url': data.get('url', "")
    })
    connection.commit()

# Función para insertar datos del clima
def insert_weather_data(data):
    cursor = connection.cursor()
    query = """
        INSERT INTO weather_data (city, temperature, feels_like, humidity, weather_description, wind_speed)
        VALUES (:city, :temperature, :feels_like, :humidity, :weather_description, :wind_speed)
    """
    weather = data['weather'][0] if 'weather' in data else {}
    cursor.execute(query, {
        'city': data.get('name'),
        'temperature': data['main']['temp'],
        'feels_like': data['main']['feels_like'],
        'humidity': data['main']['humidity'],
        'weather_description': weather.get('description'),
        'wind_speed': data['wind']['speed']
    })
    connection.commit()

# Función para obtener los datos de OpenWeather desde la base de datos
def get_weather_data_from_db():
    query = "SELECT CITY, TEMPERATURE, FEELS_LIKE, HUMIDITY, WIND_SPEED, TIMESTAMP FROM weather_data"
    cursor = connection.cursor()
    cursor.execute(query)
    rows = cursor.fetchall()
    columns = [col[0] for col in cursor.description]
    return pd.DataFrame(rows, columns=columns)

# Función para obtener los datos de NASA desde la base de datos
def get_nasa_data_from_db():
    query = "SELECT TITLE, EXPLANATION, MEDIA_TYPE, URL FROM nasa_data"
    cursor = connection.cursor()
    cursor.execute(query)
    rows = cursor.fetchall()
    columns = [col[0] for col in cursor.description]
    return pd.DataFrame(rows, columns=columns)


def generate_weather_excel():
    # Obtener los datos del clima desde la base de datos
    weather_df = get_weather_data_from_db()

    # Eliminar la columna TIMESTAMP si está presente
    if 'TIMESTAMP' in weather_df.columns:
        weather_df = weather_df.drop(columns=['TIMESTAMP'])

    # Crear el archivo Excel con la hoja de datos y las gráficas
    with pd.ExcelWriter('Clima_Graficas.xlsx', engine='openpyxl') as writer:
        # Guardar los datos de clima en la hoja de Excel
        weather_df.to_excel(writer, sheet_name='Datos de Clima', index=False)

        # Obtener el objeto de libro de trabajo de OpenPyXL
        workbook = writer.book

        # Para cada ciudad, crear una hoja con su nombre y su gráfica
        unique_cities = weather_df['CITY'].unique()

        for city in unique_cities:
            city_data = weather_df[weather_df['CITY'] == city]
            city_sheet = workbook.create_sheet(title=city)

            # Agregar los datos de la ciudad a su hoja
            for row in city_data.itertuples(index=False):
                city_sheet.append(row)

            # Crear la gráfica para la ciudad
            chart = BarChart()
            chart.title = f"Clima de {city}"
            chart.style = 10
            chart.x_axis.title = 'Indicadores'
            chart.y_axis.title = 'Valores'

            # Datos para la gráfica
            categories = ['Temperatura', 'Sensación Térmica', 'Humedad', 'Velocidad del Viento']
            values = [
                city_data['TEMPERATURE'].iloc[0],
                city_data['FEELS_LIKE'].iloc[0],
                city_data['HUMIDITY'].iloc[0],
                city_data['WIND_SPEED'].iloc[0]
            ]

            # Añadir datos a la hoja
            for i, (cat, val) in enumerate(zip(categories, values), start=2):
                city_sheet.cell(row=i, column=1, value=cat)
                city_sheet.cell(row=i, column=2, value=val)

            # Configurar los datos de la gráfica
            data_ref = Reference(city_sheet, min_col=2, min_row=2, max_row=5)
            categories_ref = Reference(city_sheet, min_col=1, min_row=2, max_row=5)
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(categories_ref)

            # Añadir colores a las barras
            fill_colors = ['FF6347', '4682B4', '32CD32', 'FFD700']  # Rojo, Azul, Verde, Amarillo
            for i, color in enumerate(fill_colors):
                if i < len(chart.series):
                    chart.series[i].graphicalProperties.solidFill = color

            # Añadir gráfica a la hoja
            city_sheet.add_chart(chart, "E5")

        # Guardar el archivo de Excel con las gráficas y datos
        workbook.save('Clima_Graficas.xlsx')


# Función para corregir el formato de fechas en el archivo Excel de NASA
def correct_nasa_dates_in_excel():
    # Obtener los datos de la NASA
    nasa_df = get_nasa_data_from_db()

    # Crear el archivo Excel con la información de la NASA, omitiendo la columna NASA_DATE
    with pd.ExcelWriter('NASA_Datos.xlsx', engine='openpyxl') as writer:
        # Guardar los datos de NASA en la hoja de Excel sin la columna 'NASA_DATE'
        nasa_df.to_excel(writer, sheet_name='Datos de NASA', index=False)

        # Obtener el objeto de libro de trabajo
        workbook = writer.book
        worksheet = workbook['Datos de NASA']

        # Guardar el archivo con la corrección de las fechas
        workbook.save('NASA_Datos.xlsx')


# Consumir mensajes e insertar en Oracle
def consume_and_store():
    print(f"Consumiendo mensajes de los topics: {', '.join(topics)}")
    last_consumed_time = time.time()  # Tiempo de la última vez que se consumió un mensaje
    max_wait_time = 10  # Tiempo máximo de espera en segundos antes de terminar
    message_count = 0  # Contador de mensajes procesados

    try:
        while True:
            msg = consumer.poll(1.0)  # Timeout de 1 segundo
            if msg is None:
                # Si no se recibe mensaje, comprobamos si ya ha pasado el tiempo máximo de espera
                if time.time() - last_consumed_time > max_wait_time:
                    print("Tiempo de espera excedido. No se recibieron más mensajes.")
                    break
                continue
            if msg.error():
                print(f"Error en el mensaje: {msg.error()}")
                continue

            topic = msg.topic()
            try:
                data = json.loads(msg.value().decode('utf-8'))
                if topic == "nasa-data":
                    insert_nasa_data(data)
                elif topic == "weather-data":
                    insert_weather_data(data)
                print(f"Mensaje insertado desde topic '{topic}': {data}")
                message_count += 1
                last_consumed_time = time.time()  # Actualizar el tiempo de consumo
            except Exception as e:
                print(f"Error procesando el mensaje de topic '{topic}': {e}")

        print(f"Se procesaron {message_count} mensajes. Cerrando consumidor.")
    except KeyboardInterrupt:
        print("Consumo detenido.")
    finally:
        # Primero generamos las gráficas y el archivo Excel de clima
        generate_weather_excel()

        # Corregir las fechas de la NASA
        correct_nasa_dates_in_excel()

        # Después cerramos el consumidor y la conexión
        consumer.close()
        connection.close()

# Main
if __name__ == "__main__":
    consume_and_store()
